# coding=utf-8
import os, sys

print(sys.getdefaultencoding())
import PostGetHttp
import platform
import json
import pymysql
from bs4 import BeautifulSoup
import http.client, urllib, urllib.request
from fdfs_client.client import *
import time
import six
os_sep = '/'



os.chdir(os.path.dirname(sys.argv[0]))


def hello_world():
    print('Hello World!')
    s = '\xe8\xb4\x9d\xe5\xa1\x94'
    s = u'我的天'
    s.encode('utf8')
    print(s)


def mtest():
    form = {'name': ['21我的天', '我的天'],
            'SpotsID': '22',
            'NameEn': 'werf',
            'NameCh': '生命',
            'Status': '1'}

    tupian = {'file': ['C:\\Users\\Public\\Pictures\\Sample Pictures\\Tulips.jpg',
                       'C:\\Users\\Public\\Pictures\\Sample Pictures\\Tulips.jpg'],
              'coverFile': 'C:\\Users\\Public\\Pictures\\Sample Pictures\\Tulips.jpg'}
    dictMerged2 = dict(form, **tupian)
    strret = PostGetHttp.posthttp(dictMerged2, 'http://localhost:8080/travel/user/test')
    print(strret)
    return json.loads(strret)


def applyCash():
    form = {'userID': '10063',
            'withdrawMoney': '200',
            'remark': 'cash',
            'userNameOfAccount': '王',
            'bankName': '招商银行',
            'accountNum': '622501254862588',
            'commission': '8',
            'passwordMD5': 'E10ADC3949BA59ABBE56E057F20F883E',
            'payUserName': '吴',
            'payBankName': '招商银行',
            'feeRate': '0.01','status': '1',
            'cashType': '2'
    }
    strret = PostGetHttp.posthttp(form, 'http://localhost:8080/travel/order/applyCash')
    return json.loads(strret)


def updateCashStatus():
    form = {'UserID': '10063',
            'cashID': '2',
            'status': '2',
            'commission': '8',
            'payAccountNum': '622825685741',
            'payUserName': '吴',
            'payBankName': '招商银行',
            'feeRate': '0.01',
            'operateUserID': '10055'
    }
    strret = PostGetHttp.posthttp(form, 'http://localhost:8080/travel/order/updateCashStatus')
    return json.loads(strret)


def mguideorderlist():
    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/guideorderlist?'
                              + 'userName=2721&mobile=0')
    return json.loads(strret)


def mgetadjustmoneylist():
    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/getadjustmoneylist?'
                              + 'userID=2721&status=0')
    return json.loads(strret)


def mgetCashList():
    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/getCashList?'
                              + 'cashID=1&status=1&guiderID=1')
    return json.loads(strret)


def mgetBankList():
    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/getBankList?'
    )
    return json.loads(strret)


def mgetWalletDetail():
    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/getWalletDetail?'
                              + 'userID=2726&page=1')
    return json.loads(strret)


def mgetBankAccountList():
    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/getBankAccountList?'
                              + 'userID=2726')
    return json.loads(strret)


def mgetWalletNumBerItem():
    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/getWalletNumBerItem?'
                              + 'userID=2726&serialNumber=201608311954115617')
    return json.loads(strret)


def mgetCashDetailByCashNo():
    strret = PostGetHttp.gethttp('10.101.1.165', '8888', '/travel/order/getCashDetailByCashNo?'
                              + 'cashNo=CA201609201739336992')
    return json.loads(strret)


def mgetWalletNumBer():
    strret = PostGetHttp.gethttp('10.101.1.165', '8888', '/travel/order/getWalletDetail?'
                              + 'userID=10063&page=1')


def mgettest():
    # form = {'name': '2721'}
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')

    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/user/gettest?'
                              + 'name=汪&page=1')
    return json.loads(strret)


def mgetMyCalendar():
    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/user/getMyCalendar?'
                              + 'guiderID=10023')
    return json.loads(strret)


def mcallRefund():
    form = {'orderID': '306',
            'orderNO': '201608311946349669',
            'batchNum': '1',
            'reason': '没有成团',
            'touristRefund': '0.01',
            'ProviderRefund': '0',
            'PlatformRefund': '0',
            'operatorID': '10063',


    }
    strret = PostGetHttp.posthttp(form, 'http://10.101.1.165:8888/apply/pay/callRefund')
    return json.loads(strret)


def msetShieldMyCalendar():
    form = {'GuideID': '10063',
    }
    strret = PostGetHttp.posthttp(form, 'http://localhost:8080/travel/user/setShieldMyCalendar')
    return json.loads(strret)


def mgetVerifyImf():
    strret = PostGetHttp.gethttp('10.101.1.165', '8888', '/travel/user/getVerifyImf?' +
                              'userID=2726'
    )
    return json.loads(strret)


def mgetbooking():
    strret = PostGetHttp.gethttp('www.booking.com', None, "/hotel/jp/the-windsor-toya-resort-spa.zh-cn.html?label=gen173nr-1FCAQoggJCC2NvdW50cnlfMTA2SCtiBW5vcmVmaDGIAQGYATLCAQNhYm7IAQzYAQHoAQH4AQOoAgQ;sid=718527c0a27d241dded2b9240b0355b8;ucfs=1;room1=A,A;dest_type=country;dest_id=106;srfid=b3812e3e0e7d34f30972e7b91c257c0fc6b91a57X13"
    )
    return json.loads(strret)

def mvisitororderdetail():
    # form = {'name': '2721'}
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')

    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/visitororderdetail?'
                              + 'orderId=571')
    return json.loads(strret)

def msearchUserByZonAndMob():
    # form = {'name': '2721'}
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')

    strret = PostGetHttp.gethttp('localhost', '8080', '/travel/order/searchUserByZonAndMob?'
                              + 'zoneCode=0086&mobile=7177')
    return json.loads(strret)

def mupdateOrderTest():
    form = {'orderid': ['2180','10064'],
            'status': ['7','1']
            }
    strret = PostGetHttp.posthttp(form, 'http://localhost:8080/travel/order/updateOrder?orderID=2180&status=7')
    return json.loads(strret)


def mgettest():
    form = {'Name': '2721',
            'Name':'wang'}
    headers = {'Content-Type':'application/x-www-form-urlencoded;charset=UTF-8',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://localhost:8080/travel/user/gettest'

    req = urllib.request.Request(url=url, data=urllib.parse.urlencode(form).encode('utf-8'), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

def mgettestjson():
    form = {'hasLanguage': '英语,日语','userName': '加油',
            'hasLanguageForshow':[{'item1':'英语','item2':'fa'},{'item1':'ri','item2':'he'}],
            'picurls':['swooefe/dkd.jpg','swodddae/dcd.jpg']}
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://localhost:8080/travel/user/gettestjson'
    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

    print(strret)
    return json.loads(strret)

def mexecuteRefund():
    strret = urllib.request.urlopen('http://localhost:8080/travel/aptest/executeRefund').read().decode('utf-8')

def mgetPlannerName():
    strret = urllib.request.urlopen('http://localhost:8080/travel/user/getPlannerName?mobileNum=13671771921').read().decode('utf-8')
    print(strret)

def mgettravellingbag():
    strret = urllib.request.urlopen('http://localhost:8080/travel/travellingbag/gettravellingbag?userID=10063').read().decode('utf-8')
    print(strret)

def mcreatePlanSchedular():
    strret = urllib.request.urlopen('http://localhost:8080/travel/travellingbag/createPlanSchedular?userID=10063').read().decode('utf-8')
    print(strret)

def msentPrivatePlanToPlanner():
    strret = urllib.request.urlopen('http://localhost:8080/travel/travellingbag/sentPrivatePlanToPlanner?userID=10063&planID=100370&guidID=10066').read().decode('utf-8')
    print(strret)

def getPlanDetail():
    strret = urllib.request.urlopen('http://10.101.1.165:8888/travel/guideplan/getPlanDetail?userID=10066&planID=100376').read().decode('utf-8')
    print(strret)

def getServicesCollectByType():
    strret = urllib.request.urlopen('http://10.101.1.165:8888/travel/travellingbag/getServicesCollectByType?userID=10063&serviceTypeID=1&page=1').read().decode('utf-8')
    print(strret)

def setFeedBack():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/user/setFeedBack?userID=10063&content=我的天空\n\nshenmedongxi\n测试中&trueName=汪先生','?&:/=')).read().decode('utf-8')
    print(strret)

def getSpotsCollectCount():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/travellingbag/getSpotsCollectCount?userID=10063','?&:/=')).read().decode('utf-8')
    print(strret)

def addOrderComplaint():
    form = {'orderID': '2548563',
            'userID': '10063',
            'iMUserID': '2500003',
            'applyRefund': '102',
            'actualRefund': '1256',
            'status': '1',
            'reason': '什么东西加油',
            'bankName': '建设银行',
            'accountNum': '6321585478521545',
            'iDNum': '34082819850122',
    }

    tupian = {'file': ['C:\\Users\\Public\\Pictures\\Sample Pictures\\Tulips.jpg',
                       'C:\\Users\\Public\\Pictures\\Sample Pictures\\Tulips.jpg']
              }

    form = dict(form, **tupian)
    strret = PostGetHttp.posthttp(form, 'http://localhost:8080/travel/order/addOrderComplaint')

    print(strret)

def getOrderComplaint():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/getOrderComplaint?complaintID=1','?&:/=')).read().decode('utf-8')
    print(strret)

def getOrderComplaint():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/getOrderComplaintID?orderID=596','?&:/=')).read().decode('utf-8')
    print(strret)

def getPlannerImfForMainPage():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/user/getPlannerImfForMainPage?userID=10063&guiderID=10067','?&:/=')).read().decode('utf-8')
    print(strret)

def getvisitororderdetail():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/visitororderdetail?orderId=565','?&:/=')).read().decode('utf-8')
    print(strret)

def getPostCode():
    from openpyxl import Workbook
    from openpyxl import load_workbook
    wb = load_workbook('D:\\workspace\\酒店booking\\tab_hotel_已更正_ - 副本.xlsx')
    ws = wb.active
    clumo = 2
    while True:
        # D列:Address
        clumoD = 'D'+str(clumo)
        # F列:Country
        clumoF = 'F'+str(clumo)
        # G列：City
        clumoG = 'G'+str(clumo)
        # H列：temppostcode
        clumoH = 'H'+str(clumo)
        # I列：postcode
        clumoI = 'I'+str(clumo)

        # C列:HotelNameCH
        clumoC = 'C'+str(clumo)
        # B列:HotelName
        clumoB = 'B'+str(clumo)

        if ws[clumoF].value == None:
            break

        if ws[clumoC].value == None:
            ws[clumoC] = ws[clumoB].value
            pass

        other = ('日本','美国')
        if ws[clumoF].value not in other:
            list = ws[clumoD].value.split(', ')
            postCode = list[len(list)-2]
            city = ws[clumoG].value
            # print('%s:%s'%(ws[clumoD].value,postCode.split(city)[0].strip()))
            ws[clumoH] = postCode.split(city)[0].strip()
            pass
        else:
            ws[clumoH] = ws[clumoI].value

        clumo = clumo+1
        pass
    wb.save('D:\\workspace\\酒店booking\\tab_hotel_已更正_ 2.xlsx')

    pass


def setAgentStatus():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/user/setAgentStatus?guideID=10079&agentGuideID=10066&status=2','?&:/=')).read().decode('utf-8')
    print(strret)

def getPrincipalGuideList():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/user/getPrincipalGuideList?guideID=10066&page=1','?&:/=')).read().decode('utf-8')
    print(strret)

def getPrincipalGuideList():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/user/getPrincipalGuideList?guideID=10066&page=1','?&:/=')).read().decode('utf-8')
    print(strret)

def getPrivatePlanDetail():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/travellingbag/getPrivatePlanDetail?planStatus=3&userID=10063&planID=100623','?&:/=')).read().decode('utf-8')
    print(strret)

def orderlistForBKM():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/orderlistForBKM?orderNo=201611','?&:/=')).read().decode('utf-8')
    print(strret)

def getCashDetail():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/getCashDetail?cashID=49','?&:/=')).read().decode('utf-8')
    print(strret)

def getCashList():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/getCashList?cashID=6','?&:/=')).read().decode('utf-8')
    print(strret)

def orderDetailForBKM():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/orderDetailForBKM?orderID=719&orderType=1','?&:/=')).read().decode('utf-8')
    print(strret)

def orderlistForBKM():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/orderlistForBKM?orderType=1&orderNo=201611251537','?&:/=')).read().decode('utf-8')
    print(strret)

def getOrderComplaintList():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/getOrderComplaintList','?&:/=')).read().decode('utf-8')
    print(strret)

def getDengniBankAccounts():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/getDengniBankAccounts','?&:/=')).read().decode('utf-8')
    print(strret)

def getFeedBackList():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/user/getFeedBackList','?&:/=')).read().decode('utf-8')
    print(strret)

def getFeedBackDetail():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/user/getFeedBackDetail?feedbackID=1','?&:/=')).read().decode('utf-8')
    print(strret)

def saveDeviceTokenToServer():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/auth/saveDeviceTokenToServer?userID=10063&deviceToken=a72bf8b8dc87d8107db961061ba2fcc00735419ae66c3e841d33599f5734ab00','?&:/=')).read().decode('utf-8')
    print(strret)

def guideorderdetail():
    strret = urllib.request.urlopen(urllib.parse.quote('http://10.101.1.165:8888/travel/order/guideorderdetail?orderId=743','?&:/=')).read().decode('utf-8')
    print(strret)



def getAgent():
    strret = urllib.request.urlopen(urllib.parse.quote('http://10.101.1.165:8888/travel/user/getAgent?guideID=10023','?&:/=')).read().decode('utf-8')
    print(strret)

def changeOrderPrice():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/changeOrderPrice?orderID=810&price=1','?&:/=')).read().decode('utf-8')
    print(strret)

#json传值 post，添加更新件数的通知消息
def putMessage():
    headerdata = {"Host":"10.101.1.36",'Content-Type':'application/json;charset=UTF-8'}
    #json实体
    test_data = {'relativeType':'31','relativeID':'5','secondType':1}
    # json 编码
    test_data_urlencode = json.JSONEncoder().encode(test_data)
    conn = http.client.HTTPConnection("10.101.1.36",8080)
    requrl = '/travel/platformservice/putMessage'

    conn.request(method="POST",url=requrl,body=test_data_urlencode,headers = headerdata)

    response = conn.getresponse()
    res= response.read()
    print(res)

# 查询当前消息列表各类型和条数
def getAllMessageTypeCount():
    strret = urllib.request.urlopen(urllib.parse.quote('http://10.101.1.36:8080/travel/platformservice/getAllMessageTypeCount','?&:/=')).read().decode('utf-8')
    print(strret)

def getwxpayinfo():
    form = {"appid":"wxce4e9a43fa69a9ce","mchId":"1485441592","objType":"0","outTradeNo":"201707181816353320","payerID":"10087","totalFee":"200"}
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://10.101.1.165:8096/apply/pay/getwxpayinfo'
    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

    print(strret)
    return json.loads(strret)


def createGroup():
    form = {"groupName":"GroupTest","createUserID":"10063","userGroupInfos":[{"userID":2721},{"userID":2726},{"userID":10058},{"userID":10087}]}
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://localhost:8080/travel/im/createGroup'
    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

    print(strret)
    return json.loads(strret)

# 查询当前消息列表各类型和条数
def getGroupUsers():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/im/getGroupUsers?groupID=10&userID=10063','?&:/=')).read().decode('utf-8')
    print(strret)

# /travel/im/groupRefresh
def groupRefresh():
    form = {"groupName":"GroupTestRE","groupID":"8","userGroupInfos":[{"userID":10063},{"userID":10058},{"userID":10087}]}
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://localhost:8080/travel/im/groupRefresh'
    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

    print(strret)
    return json.loads(strret)

def deleteGroupUsers():
    form = {"groupID":"28","userGroupInfos":[{"userID":2721},{"userID":2726}]}
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://localhost:8080/travel/im/deleteGroupUsers'
    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

    print(strret)
    return json.loads(strret)

# gName,groupID,userID
def groupRefreshSelfGname():
    strret = urllib.request.urlopen(urllib.parse.quote('http://10.101.1.165:8096/travel/im/groupRefreshSelfGname?groupID=53&userID=2726&gName=哈哈','?&:/=')).read().decode('utf-8')
    print(strret)

#     groupQuit
#  {groupID,userID}
def groupQuit():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/im/groupQuit?groupID=8&userID=10058','?&:/=')).read().decode('utf-8')
    print(strret)

#     /getGroupPic
def getGroupPic():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/im/getGroupPic?groupID=10','?&:/=')).read().decode('utf-8')
    print(strret)

#     /setUserNameVisible
def setUserNameVisible():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/im/setUserNameVisible?groupID=10&userID=10058&userNameVisibleFlag=1','?&:/=')).read().decode('utf-8')
    print(strret)

# /getGroupInfoForSelf
def getGroupInfoForSelf():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/im/getGroupInfoForSelf?groupID=10&userID=10058','?&:/=')).read().decode('utf-8')
    print(strret)

# /getGroupInfoForSelf
def getGroupInfoListForSelf():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/im/getGroupInfoListForSelf?userID=10058','?&:/=')).read().decode('utf-8')
    print(strret)

# /getGroupInfoForSelf
def setFriendMemoName():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/im/setFriendMemoName?fromUserID=10063&toUserID=10087&memoName=设置者10063','?&:/=')).read().decode('utf-8')
    print(strret)

    # /getFriendMemoName
def getFriendMemoName():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/im/getFriendMemoName?fromUserID=10063&toUserID=10087&memoName=设置者10063','?&:/=')).read().decode('utf-8')
    print(strret)

def maopao():

    ints = [3,4,6,9,11,3,5,16,2,8]
    # dict = {'Alice': '2341', 'Beth': '9102', 'Cecil': '3258'}
    # print(len(dict))
    # print(dict.__len__())
    # print(len(ints))
    # print(ints.__len__())
    for j, val2 in enumerate(ints) :

        for i,val in enumerate(ints):
            if i == len(ints)-2-j:
                break
            if ints[i] > ints[i+1]:
                temp = ints[i+1]
                ints[i+1] = ints[i]
                ints[i] = temp


        print('index %d val %d'%(i,val))
        pass

    pass


def joinGroup():
    form = {"groupID":"56","createUserID":"10000202","userGroupInfos":[{"userID":2721},{"userID":2726}]}
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://localhost:8080/travel/im/joinGroup'
    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

    print(strret)
    return json.loads(strret)

def short_url():
    params = {"source":267208165}

    urlLongs = ["http://issue.dengnilvyou.com.cn/registerIndex?inviteCode=99999","http://issue.dengnilvyou.com.cn/registerIndex?inviteCode=15421"]
    for i,urllong in enumerate(urlLongs):

        key = 'url_long'
        params[key] = urllong
        pass

    urllong  =  urllib.parse.urlencode(params)
    strret = urllib.request.urlopen(urllib.parse.quote('http://api.t.sina.com.cn/short_url/shorten.json?','?&:/=')+urllong).read().decode('utf-8')
    print(strret)

def applyRefund():
    form = {"applyUserID":"10000202","orderID":"12584","refundMoney":222.3,"refundReason":"测试ga"}
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://localhost:8080/travel/order/applyRefund'
    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

    print(strret)
    return json.loads(strret)

    # /getIMRelationApplyListExtra
def getIMRelationApplyListExtra():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/userrm/getIMRelationApplyListExtra?userID=10063','?&:/=')).read().decode('utf-8')
    print(strret)


    # * @param  redPacket {money 紅包金額,whoGetType 1 全员 2 指定人  单聊红包无需指定,
    # *                   redPacketType 类型 1普通，2拼手气，3单聊,number 红包数量,
    # *                   blessing 红包祝福语,userID 发红包的用户,toID 可以是群 可以是单聊的对方userID
    # *                   redPacketMemberList 指定人红包时的user:[userID:10063,userID:10053]
    # *                   }
def createRedPacket():
    form = {'money': '10',
            'whoGetType':'1',
            'redPacketType':'2',
            'number':'3',
            'blessing':'wang',
            'userID':'10063',
            'toID':'77'
        # ,
        #     'redPacketMemberList':[{'userID':10063},{'userID':2721},{'userID':2726}]
    }
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    url = 'http://localhost:8080/travel/order/createRedPacket'

    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    print(strret)
    pass

def getRedPacketMoney():
    # 10053 10063 10000062
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/getRedPacketMoney?userID=10000161&redPacketID=20','?&:/=')).read().decode('utf-8')
    print(strret)

def addVirtualOrder():

    # @param  {consumeType 消费种类:1，行程打赏，2，加入方案, 3,创建自定义需求, 4,加入方案, 5,门票转让, 6,收付款, 7,红包
    #                                    *                    serviceID 商品ID（接受转让门票时，是转让门票的 transferOrderID，转账,加入红包时,填创建的红包的ID ）, serviceName 名字,payType 支付方式(4，钱包,1,支付宝，2,银行卡,3,微信)
    # *                    money 金额, sellerID 卖方ID（转账时，为收款方，创建红包时，不设置）, buyerID 买方ID（转账时，为付款方，创建红包时，为付款方）,passWordMD5:MD5密码}
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/order/addVirtualOrder?consumeType=7&serviceID=8&serviceName=%e7%ba%a2%e5%8c%85&payType=1&money=2721&buyerID=10063','?&:/=')).read().decode('utf-8')
    print(strret)

def pushConsultMessage():
    strret = urllib.request.urlopen(urllib.parse.quote('http://localhost:8080/travel/platformservice/pushConsultMessage?fromUserID=1099&toUserID=10063','?&:/=')).read().decode('utf-8')
    print(strret)

def jsonCommon(url,form):
    headers = {'Content-Type':'application/json',
               "Accept": "application/json;charset=UTF-8",
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'}
    form = json.dumps(form)
    req = urllib.request.Request(url=url, data=form.encode("utf-8"), headers=headers,method='POST')
    strret = urllib.request.urlopen(req).read().decode('utf-8')
    # strret = PostGetHttp.posthttp(form,'http://localhost:8080/travel/user/gettest')
    pass

    print(strret)
    return json.loads(strret)

def deleteCardTips():
    form = {"userID":"10063","touristCardTipIDs":[15868,36987]}
    url = 'http://localhost:8080/travel/touristCard/deleteCardTips'

    ret = jsonCommon(url,form)

def refundcard():
    form = {"passwordMD5":"E10ADC3949BA59ABBE56E057F20F883E","userID":"10063","userCardIDs":[15868,36987]}
    url = 'http://localhost:8080/travel/order/refundcard'

    ret = jsonCommon(url,form)

if __name__ == '__main__':
    # deleteCardTips()
    refundcard()

    # pushConsultMessage()
    # mupdateOrderTest()
    # getRedPacketMoney()
    # addVirtualOrder()
    # createRedPacket()
    # getIMRelationApplyListExtra();
    # applyRefund()
    # mgettest()
    # short_url()
    # deleteGroupUsers()
    # joinGroup()
    # maopao()
    # getFriendMemoName()
    # setFriendMemoName()
    # getGroupInfoListForSelf()
    # setUserNameVisible()
    # getGroupPic()
    # groupQuit()
    # groupRefreshSelfGname()
    # groupRefresh()
    # getGroupUsers()
    # createGroup()
    # sentPrivatePlanToPlanner
    # mcreatePlanSchedular()
    # msentPrivatePlanToPlanner()
    # getwxpayinfo()
    # mgettestjson()
    # getAllMessageTypeCount()
    # putMessage()
    # addOrderComplaint()
    # getPlanDetail()
    # getAgent()

    # getServicesCollectByType()
    # setFeedBack()

    # addOrderComplaint()

    # getOrderComplaint()

    # getPlannerImfForMainPage()

    # getvisitororderdetail()

    # getPostCode()

    # setAgentStatus()

    # getPrincipalGuideList()

    # getPrivatePlanDetail()

    # orderlistForBKM()
    # getCashDetail()
    # getCashList()

    # orderDetailForBKM()
    # orderlistForBKM()

    # getOrderComplaintList()

    # getFeedBackList()
    # guideorderdetail()

    # getFeedBackDetail()

    # applyCash()
    # mupdateOrderTest()
    # getDengniBankAccounts()
    # mgettest()

    # saveDeviceTokenToServer()
    # changeOrderPrice()
    # jsonobj = mtest()
    exit()
    # PostGetHttp.gethttp('localhost','8080','/travel/order/orderDetailForBKM?'
    # +'orderID=324&orderType=2')
    # PostGetHttp.gethttp('10.101.1.165','8888','/travel/order/visitororderlist?'
    # +'userID=2721&status=0&page=1')
    #


    # jsonobj= mguideorderlist()

    # jsonobj = mgetbooking()
    # print(jsonobj)


    # conn = pymysql.connect(host='10.101.1.163', port=3306, user='traveldb', passwd='traveldb', db='traveldb',
    #                        charset='UTF8')
    # cur = conn.cursor()
    # cur.execute("select version()")
    # conn.commit()
    # for i in cur:
    #     print(i)
    # cur.close()
    # conn.close()
    #
    # hotel_picscr='http://r-cc.bstatic.com/images/hotel/square200/139/13918842.jpg'
    #
    # cover = urllib.request.urlopen(hotel_picscr)
    # pic_name = hotel_picscr.split('/').pop()
    # path = 'D:/Download'+'/'+'123'+'/'+'2321'+'/'
    # os.makedirs(path,exist_ok=True)
    # f = open(path+pic_name,'wb+')
    # f.write(cover.read())
    # f.flush()
    # f.close()
    # cover.close()

    hotel_picscr = '222840x4603333'
    hotel_picscr.replace('840x460','max1024x768')
    hotel_picscr.replace('max1024x768','max400')

    send_buffer = struct.pack('!Q Q c 16s', 44,21,b'0',b'group1')

    send_buffer = struct.pack('!Q Q c 16s 44s 21s', 44,21,b'0',b'group1',
                              b'M00/06/E8/CmUBpVf_Q76AS3jhAABO9Nigqbg929.jpg',
                              b'fileName43062953.jpg')


    file = urllib.request.urlopen('http://r-cc.bstatic.com/images/hotel/max400/430/43062953.jpg')

    client_file='fdfs_client.conf'
    test_file='test.txt'
    download_file='test2.txt'

    try:
        client = Fdfs_client(client_file)
        #upload

        ret_upload = client.upload_by_buffer(file.read(),'jpg',{"fileName":"43062953.jpg"})
        print (ret_upload)

        time.sleep(1)   #等待5s，否则下载时会报错文件不存在
        file.close()
        file_id=ret_upload['Remote file_id'].replace('\\','/')  #新版本文件存放Remote file_id格式变化

        #download
        # ret_download=client.download_to_file(download_file,file_id)
        # print (ret_download)

        #delete
        ret_delete=client.delete_file(file_id)

        print (ret_delete)

    except Exception as ex:
        print (ex)

    # path = 'D:/Download'+'/'+'123'+'/'+'45'+'/'
    # os.makedirs(path,exist_ok=True)
    # f = open(path+'1.jpg','wb')
    #
    # f.close()
    # 模拟chrome登陆
    # headers = {
    #     'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.106 Safari/537.36',
    #     'Accept': 'text/html;q=0.9,*/*;q=0.8', 'Accept-Charset': 'utf-8;q=0.7,*;q=0.3',
    #     'Connection': 'close'}
    #
    # opener = urllib.request.build_opener()
    # opener.addheaders = headers.items()
    #
    # hotelHtml = opener.open(
    #     'http://www.booking.com/hotel/jp/keio-plaza-tokyo.zh-cn.html?label=gen173nr-1FCAQoggJCC2NvdW50cnlfMTA2SCtiBW5vcmVmaDGIAQGYATLCAQNhYm7IAQzYAQHoAQH4AQOoAgQ;sid=b23ba36a407c172ed759586ddb142cea;ucfs=1;room1=A,A;dest_type=country;dest_id=106;srfid=b3812e3e0e7d34f30972e7b91c257c0fc6b91a57X3').read().decode(
    #     'utf-8')
    # hotelHtmlSoup = BeautifulSoup(hotelHtml, "html.parser")
    #
    # # 地址
    # hotel_address = hotelHtmlSoup.find(itemprop="address").string
    # photos_distinct = hotelHtmlSoup.find(id='photos_distinct')
    # if None != photos_distinct:
    #     photos_a_ss = photos_distinct.find_all('a')
    # elif None == photos_distinct:
    #     photos_a_ss = hotelHtmlSoup.find_all('a', class_='bh-photo-grid-item')
    # photo_dict = {}
    # for photos_a in photos_a_ss:
    #
    #     if photos_a.has_attr('data-photoid'):
    #         photo_id = photos_a['data-photoid']
    #     elif photos_a.has_attr('data-id'):
    #         photo_id = photos_a['data-id']
    #
    #     photo_dict[photo_id] = photos_a['href']
    #
    # hotel_desc_summary = hotelHtmlSoup.find(id='summary')
    # hotel_desc_main=hotel_desc_summary.get_text()
    #
    # deleteitme = hotel_desc_summary.find('span',class_='city_centre_map_link')
    # if(None != deleteitme):
    #     # 移除文档树
    #     deleteitme.decompose()
    #
    # deleteitme = hotel_desc_summary.find('span',class_='city_centre_map_link')

    # PostGetHttp.gethttp('10.101.1.165','8888','/travel/order/visitororderlist?'
    #         +'userID=2721&status=0')
    # hello_world()
# hello_world()
# print ('"Hello World!"')
#     a=3
#     if a==2:
#         print('2')
#     elif a==3:
#         print('3')
#     a=5
#     while a>0:
#         a=a-1
#         print(a)
#     for i  in  range(6,10):
#         print(i)
#     array = {'China',
#              'America','England'}
#
#     for aitem  in  array:
#         print(aitem)
#
#     f=open('Python.py','r+')
#     print(f.readlines())
#     path1=os.getcwd()
#     path2=sys.argv[0]
#     print(path2)
#
#     name = input("Name:\n")
#     print(name)













